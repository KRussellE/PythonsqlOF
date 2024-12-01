import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from PIL import Image, ImageTk  # Import image handling for icons
import shutil
import sqlite3
import os  # For file existence check

# Tkinter window creation
root = tk.Tk()
root.title("File Opener and SQL Connection")

# Window size
root.geometry("1000x800")  # Enlarged window size

# Set the background color for the window
root.configure(bg="#f0f0f0")  # Light gray background

# Color scheme
primary_color = "#4A90E2"  # Blue color
secondary_color = "#D1D3D4"  # Light gray color
highlight_color = "#E74C3C"  # Red color
button_color = "#3498DB"  # Blue for buttons
button_hover_color = "#2980B9"  # Darker blue for hover effect
text_color = "#2C3E50"  # Dark gray text

# Load icons for buttons
def load_icon(icon_name, size=(30, 30)):
    """Load icon image and resize it."""
    try:
        icon = Image.open(f"{icon_name}.png")  # Ensure icons are in the same folder as your script
        icon = icon.resize(size, Image.ANTIALIAS)
        return ImageTk.PhotoImage(icon)
    except Exception as e:
        print(f"Error loading icon {icon_name}: {e}")
        return None

# Load icons
open_file_icon = load_icon("open_file")  # You need a "open_file.png" icon
connect_sql_icon = load_icon("connect_sql")  # You need a "connect_sql.png" icon
update_icon = load_icon("update")  # You need a "update.png" icon
download_db_icon = load_icon("download_db")  # You need a "download_db.png" icon

# Text box for file output and updates
output_box = tk.Text(root, wrap=tk.WORD, height=15, width=70, bg="white", fg=text_color, font=("Arial", 12))
output_box.grid(row=1, column=0, columnspan=2, pady=20)

# SQL connection status label
connection_status_label = tk.Label(root, text="", fg=primary_color, font=("Helvetica", 12), bg="#f0f0f0")
connection_status_label.grid(row=2, column=0, columnspan=2, pady=5)

# Store database connection status in root
root.db_connected = False

# Open File function
def open_file():
    file_path = filedialog.askopenfilename(title="Select an Excel file", filetypes=[("Excel Files", "*.xlsx;*.xls")])
    if file_path:
        try:
            book = load_workbook(file_path)
            sheet = book['Sheet1']
            keys = [sheet.cell(row=1, column=col_index + 1).value for col_index in range(sheet.max_column)]
            
            dict_list = []
            for row_index in range(2, sheet.max_row + 1):
                d = {keys[col_index]: sheet.cell(row=row_index, column=col_index + 1).value
                     for col_index in range(sheet.max_column)}
                dict_list.append(d)
            
            output_box.delete(1.0, tk.END)
            output_box.insert(tk.END, f"File loaded: {file_path}\n\n")
            for d in dict_list:
                output_box.insert(tk.END, f"{d}\n")
            # Save the data for later use
            root.excel_data = dict_list  # Save data in root object

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while loading the file: {e}")
    else:
        messagebox.showwarning("Warning", "No file was selected.")

def connect_to_sql():
    """Connect to the SQL database."""
    try:
        # Assume the SQL connection logic is in sqlconnection.py
        conn = sqlconnection.connect_to_database()  # This function should handle the connection
        if conn:
            root.db_connected = True  # Set the database connected flag
            connection_status_label.config(text="Successfully connected to SQL", fg="green", bg="#f0f0f0")
        else:
            root.db_connected = False  # Set the database connected flag to False
            connection_status_label.config(text="SQL connection failed", fg="red", bg="#f0f0f0")
    except Exception as err:
        messagebox.showerror("SQL Error", f"An error occurred during connection: {err}")
        connection_status_label.config(text="SQL connection failed", fg="red", bg="#f0f0f0")

# Function to download database with overwrite confirmation
def download_database():
    """Download the SQLite database to a file."""
    if not root.db_connected:
        messagebox.showwarning("Error", "No connection to the database!")
        return

    # Ask the user for the location to save the database file
    file_path = filedialog.asksaveasfilename(defaultextension=".db", filetypes=[("SQLite DB Files", "*.db")])
    
    if file_path:
        # Check if the file already exists
        if os.path.exists(file_path):
            # Ask the user if they want to overwrite the existing file
            overwrite = messagebox.askyesno("Overwrite File", "The database file already exists. Do you want to overwrite it?")
            if not overwrite:
                messagebox.showinfo("Cancelled", "The download was cancelled.")
                return  # If user chooses not to overwrite, exit the function

        try:
            # Assuming the connection is to a SQLite database, we will simply copy the DB file
            # Example path to the existing SQLite database file (this will depend on your system)
            existing_db_file = 'path_to_your_existing_sqlite_db_file.db'

            # Copy the SQLite database file to the chosen location
            shutil.copy(existing_db_file, file_path)
            messagebox.showinfo("Success", f"Database has been successfully downloaded to: {file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while downloading the database: {e}")

def update_data():
    """Call the update function and show the results."""
    # Check if database is connected
    if not root.db_connected:
        # Show error message if not connected to the database
        update_result_box.configure(state=tk.NORMAL)
        update_result_box.delete(1.0, tk.END)
        update_result_box.insert(tk.END, "ERROR: No database connection.\nPlease connect to the database first.\n")
        update_result_box.configure(state=tk.DISABLED)
        return
    
    if hasattr(root, 'excel_data'):
        # Call the update function and get the results
        try:
            # Call the update function that will use the SQL connection to update data
            results, found_count, not_found_count = sqlconnection.update_data_from_sql(root.excel_data)  # Replace with actual method from sqlconnection.py
            if results:
                update_result_box.configure(state=tk.NORMAL)  # Enable the result box to insert text
                update_result_box.delete(1.0, tk.END)  # Clear any previous content
                update_result_box.insert(tk.END, f"Update Results:\n{results}")
                update_result_box.insert(tk.END, f"\nFound {found_count} Docket Numbers, {not_found_count} not found.")
                update_result_box.configure(state=tk.DISABLED)  # Disable it again after updating
            else:
                update_result_box.configure(state=tk.NORMAL)
                update_result_box.delete(1.0, tk.END)
                update_result_box.insert(tk.END, "ERROR: No data was updated.\n")
                update_result_box.configure(state=tk.DISABLED)

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during data update: {e}")
            update_result_box.configure(state=tk.NORMAL)
            update_result_box.delete(1.0, tk.END)
            update_result_box.insert(tk.END, f"ERROR: {e}")
            update_result_box.configure(state=tk.DISABLED)
    else:
        messagebox.showwarning("Warning", "No data loaded from the Excel file.")


# Create widgets for buttons and text boxes using grid layout
open_file_button = tk.Button(root, text="Open File", command=open_file, bg=button_color, fg="white", font=("Arial", 12), relief="flat", image=open_file_icon, compound="left")
open_file_button.grid(row=0, column=0, pady=10, padx=20)

connect_sql_button = tk.Button(root, text="Connect to SQL", command=connect_to_sql, bg="#006400",  # Dark green background
    fg="white",    # White text color
    font=("Arial", 12), relief="flat", image=connect_sql_icon, compound="left")
connect_sql_button.grid(row=3, column=1, pady=10, padx=20)

download_db_button = tk.Button(root, text="Download Database", command=download_database, bg="#32CD32",  # Lighter green background
    fg="white",    # White text color
    font=("Arial", 12), relief="flat", image=download_db_icon, compound="left")
download_db_button.grid(row=4, column=1, pady=10, padx=20)

update_data_button = tk.Button(root, text="Update Data", command=update_data, bg="#F39C12", fg="white", font=("Arial", 12), relief="flat", image=update_icon, compound="left")
update_data_button.grid(row=5, column=1, pady=10, padx=20)

# Create a result text box for update results
update_result_box = tk.Text(root, height=10, width=70, bg="white", fg=text_color, font=("Arial", 12))
update_result_box.grid(row=6, column=0, columnspan=2, pady=20)
update_result_box.configure(state=tk.DISABLED)  # Disable it initially

# Run the GUI application
root.mainloop()
