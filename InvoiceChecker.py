import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog  # Import simpledialog for password input
from openpyxl import load_workbook
from PIL import Image, ImageTk  # Import image handling for icons
import shutil
import sqlite3
import os  # For file existence check
import pymysql
from pymysql import Error
from PIL import Image
import subprocess
import threading

# Tkinter window creation
root = tk.Tk()
root.title("File Opener and SQL Connection")

# Window size
root.geometry("1200x900")  # Enlarged window size

# Set the background color for the window
root.configure(bg="#f0f0f0")  # Light gray background

# Color scheme
primary_color = "#4A90E2"  # Blue color
secondary_color = "#D1D3D4"  # Light gray color
highlight_color = "#E74C3C"  # Red color
button_color = "#3498DB"  # Blue for buttons
button_hover_color = "#2980B9"  # Darker blue for hover effect
text_color = "#2C3E50"  # Dark gray text

# Label a letöltés státuszának megjelenítéséhez
status_label = tk.Label(root, text="", fg="black", font=("Helvetica", 12), bg="#f0f0f0")
status_label.grid(row=4, column=0, columnspan=2, pady=10)

# Load icons for buttons
def load_icon(icon_name, size=(30, 30)):
    """Load icon image and resize it."""
    try:
        icon = Image.open(f"{icon_name}.png")  # Ensure icons are in the same folder as your script
        icon = icon.resize(size, Image.ANTIALIAS)
        return ImageTk.PhotoImage(icon)
    except Exception as e:
        print(f"Error loading icon {icon_name}: {e}")
        return None

# Load icons
open_file_icon = load_icon("open_file")  # You need a "open_file.png" icon
connect_sql_icon = load_icon("connect_sql")  # You need a "connect_sql.png" icon
update_icon = load_icon("update")  # You need a "update.png" icon
download_db_icon = load_icon("download_db")  # You need a "download_db.png" icon

# Text box for file output and updates
output_box = tk.Text(root, wrap=tk.WORD, height=15, width=70, bg="white", fg=text_color, font=("Arial", 12))
output_box.grid(row=1, column=0, columnspan=2, pady=20)

# SQL connection status label
connection_status_label = tk.Label(root, text="", fg=primary_color, font=("Helvetica", 12), bg="#f0f0f0")
connection_status_label.grid(row=2, column=0, columnspan=2, pady=5)

# Store database connection status in root
root.db_connected = False

def open_file():
    file_path = filedialog.askopenfilename(title="Select an Excel file", filetypes=[("Excel Files", "*.xlsx;*.xls")])
    if file_path:
        try:
            book = load_workbook(file_path)
            sheet = book['Sheet1']
            keys = [sheet.cell(row=1, column=col_index + 1).value for col_index in range(sheet.max_column)]
            
            dict_list = []
            item_barcodes = []  # Lista a barcode-ok tárolására
            for row_index in range(2, sheet.max_row + 1):
                d = {keys[col_index]: sheet.cell(row=row_index, column=col_index + 1).value
                     for col_index in range(sheet.max_column)}
                dict_list.append(d)
                item_barcodes.append(d['Item Barcode'])  # Adja hozzá a barcode-ot a listához
            
            output_box.delete(1.0, tk.END)
            output_box.insert(tk.END, f"File loaded: {file_path}\n\n")
            for d in dict_list:
                output_box.insert(tk.END, f"{d}\n")
            # Store the barcode list for later use
            root.item_barcodes = item_barcodes  # Store the item barcodes in root object
            root.excel_data = dict_list  # Save data in root object

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while loading the file: {e}")
    else:
        messagebox.showwarning("Warning", "No file was selected.")

def connect_to_sql(output_box, status_label):
    try:
        # Ask the user for the password using a simple dialog
        password = simpledialog.askstring("Password", "Please enter your SQL password:", show='*')
        if not password:
            messagebox.showwarning("Warning", "Password is required to connect.")
            return

        # Kapcsolódás az adatbázishoz
        connection = pymysql.connect(
            host="access-sync.cnomqm8qwozn.eu-north-1.rds.amazonaws.com",
            user="Ogden",
            password=password,  # Use the entered password
            database="Access-Info"
        )

        # Try to execute a simple query to check if connection is successful
        cursor = connection.cursor()
        cursor.execute("SELECT 1")
        cursor.close()

        # If no exception was raised, the connection is successful
        output_box.insert(tk.END, "\n\nSQL kapcsolat sikeresen létrejött!\n")

        # Now check available tables
        cursor = connection.cursor()
        cursor.execute("SHOW TABLES")
        tables = cursor.fetchall()

        output_box.insert(tk.END, "Táblák az adatbázisban:\n")
        for table in tables:
            output_box.insert(tk.END, f"{table[0]}\n")

        cursor.close()
        connection.close()

        # Update connection status label
        status_label.config(text="Sikeresen csatlakozott az SQL adatbázishoz", fg="green")
        root.db_connected = True  # Set flag to indicate successful connection

    except pymysql.Error as e:
        output_box.insert(tk.END, f"Hiba az SQL kapcsolódás során: {e}\n")
        status_label.config(text="SQL kapcsolat sikertelen", fg="red")
        root.db_connected = False  # Set flag to indicate failed connection


# Add a label for showing the loading status
loading_status_label = tk.Label(root, text="", fg="black", font=("Helvetica", 12), bg="#f0f0f0")
loading_status_label.grid(row=3, column=0, padx=10, pady=10)  # Place it next to the button

# Add a new Text widget below the "Download Database" button to display query results
query_output_box = tk.Text(root, wrap=tk.WORD, height=10, width=80, bg="white", fg=text_color, font=("Arial", 12))
query_output_box.grid(row=4, column=0, columnspan=2, pady=20)

# Szövegmező a státusz üzenetekhez
loading_status_label = tk.Label(root, text="Ready", font=("Arial", 12), fg="green")
loading_status_label.grid(row=0, column=1, pady=10)

def run_sql_query(password):
    # A jelszóval és a barcodes listával kapcsolódunk az SQL adatbázishoz

    # A lekérdezés előtt a státuszt frissítjük
    loading_status_label.config(text="Executing query... 0%", fg="black")
    
    barcodes = root.item_barcodes
    if not barcodes:
        loading_status_label.config(text="No barcodes found", fg="red")
        messagebox.showwarning("Warning", "No barcodes loaded from the Excel file!")
        return

    try:
        # SQL kapcsolat létrehozása
        connection = pymysql.connect(
            host="access-sync.cnomqm8qwozn.eu-north-1.rds.amazonaws.com",
            user="Ogden",
            password=password,
            database="Access-Info"
        )
        cursor = connection.cursor()

        query_output_box.delete(1.0, tk.END)  # Töröljük az előző tartalmat
        query_output_box.insert(tk.END, "Executing query for the following barcodes:\n")
        query_output_box.insert(tk.END, "\n".join(barcodes) + "\n\n")

        # Az oszlopok, amelyeket szeretnénk kiírni
        columns_of_interest = [
            "ID", "Client Id", "Warehouse Id", "Order Number", "External Order Reference",
            "Order Date", "Despatch Date", "Total Weight", "Tracking Number",
            "Client Short Name", "Client Name", "Client Code", "Order Status Name",
            "Warehouse Name", "OrderNumber+ClientId"
        ]

        # SQL lekérdezés végrehajtása minden barcode-ra
        total_barcodes = len(barcodes)
        for i, barcode in enumerate(barcodes):
            query = f"""
                SELECT 
                    `ID`, `Client Id`, `Warehouse Id`, `Order Number`, `External Order Reference`,
                    `Order Date`, `Despatch Date`, `Total Weight`, `Tracking Number`,
                    `Client Short Name`, `Client Name`, `Client Code`, `Order Status Name`,
                    `Warehouse Name`, `OrderNumber+ClientId`
                FROM `_Orders-FinalView-v02_sync`
                WHERE `Tracking number` = %s
            """
            cursor.execute(query, (barcode,))
            results = cursor.fetchall()

            # Frissítjük a státuszt a feldolgozás előrehaladásával
            progress_percent = int(((i + 1) / total_barcodes) * 100)  # Százalékos előrehaladás
            loading_status_label.config(text=f"Executing query... {progress_percent}%")
            loading_status_label.update()  # Frissítjük az UI-t

            query_output_box.insert(tk.END, f"Executed query for barcode: {barcode}\n")
            if results:
                query_output_box.insert(tk.END, f"Results for barcode {barcode}:\n")
                for result in results:
                    # Eredmény dictionary-ben történő kiírása
                    result_dict = {columns_of_interest[i]: result[i] for i in range(len(columns_of_interest))}
                    query_output_box.insert(tk.END, f"{result_dict}\n")
            else:
                query_output_box.insert(tk.END, f"No results for barcode {barcode}\n")

        cursor.close()
        connection.close()

        # Státusz frissítése "Ready"-ra
        loading_status_label.config(text="Ready", fg="green")

    except pymysql.MySQLError as e:
        loading_status_label.config(text="SQL query failed", fg="red")
        query_output_box.insert(tk.END, f"Error during SQL query: {e}\n")

# Button to execute SQL query in a separate thread
def execute_sql_query_thread():
    thread = threading.Thread(target=execute_sql_query)
    thread.start()

# Function to handle matching of tracking numbers
def match_tracking_numbers():
    if not hasattr(root, 'item_barcodes') or len(root.item_barcodes) == 0:
        messagebox.showwarning("Warning", "No item barcodes loaded from Excel file!")
        return

    # Clear the right output box before displaying new results
    right_output_box.delete(1.0, tk.END)
    right_output_box.insert(tk.END, "Matching Tracking Numbers:\n")

    # Dictionary to store the combined data
    combined_data = {}

    # Iterate through all item barcodes
    for barcode in root.item_barcodes:
        # Find the matching item from the Excel data
        matching_item = None
        for item in root.excel_data:
            if item['Item Barcode'] == barcode:
                matching_item = item
                break

        if matching_item:
            # Now query the SQL database for the barcode
            query = f"""
            SELECT `ID`, `Client Id`, `Warehouse Id`, `Order Number`, `Total Weight`, `Order Value`,
            `Tracking Number`, `Client Short Name`, `Client Name`, `Client Code`, `Order Status Name`
            FROM `_Orders-FinalView-v02_sync`
            WHERE `Tracking Number` = '{barcode}'
            """
            
            try:
                # Open the SQL connection again
                connection = pymysql.connect(
                    host="access-sync.cnomqm8qwozn.eu-north-1.rds.amazonaws.com",
                    user="Ogden",
                    password="wLzp7ueqgGigbzL",
                    database="Access-Info"
                )
                cursor = connection.cursor()

                cursor.execute(query)
                rows = cursor.fetchall()

                if rows:
                    # We assume only one row for the barcode, we take the first row (if found)
                    sql_data = dict(zip([column[0] for column in cursor.description], rows[0]))

                    # Combine the Excel data and SQL data into a single dictionary
                    combined_item = {**matching_item, **sql_data}
                    combined_data[barcode] = combined_item

                    # Insert combined data into the right output box in dictionary format
                    right_output_box.insert(tk.END, f"{combined_item}\n\n")
                else:
                    # If no data was found in the SQL database for the barcode
                    right_output_box.insert(tk.END, f"No data found for barcode {barcode} in SQL.\n\n")

                # Close the SQL connection
                cursor.close()
                connection.close()

            except pymysql.Error as e:
                right_output_box.insert(tk.END, f"Error querying the database for barcode {barcode}: {e}\n\n")

        else:
            right_output_box.insert(tk.END, f"No matching Excel data for barcode {barcode}.\n\n")

    # Display completion message
    right_output_box.insert(tk.END, "\nMatching complete!")


# Text box for displaying data on the right side of the window
right_output_box = tk.Text(root, wrap=tk.WORD, height=15, width=40, bg="white", fg=text_color, font=("Arial", 12))
right_output_box.grid(row=1, column=2, rowspan=4, padx=20, pady=20)  # Right side placement

# Optional: Add a label above the right output box
right_output_label = tk.Label(root, text="Right Output Box", fg="black", font=("Helvetica", 12), bg="#f0f0f0")
right_output_label.grid(row=0, column=2, pady=5)

# Now we can use the right_output_box to display data
def show_data_in_right_box(data):
    """Function to insert data into the right output box."""
    right_output_box.delete(1.0, tk.END)  # Clear previous content
    right_output_box.insert(tk.END, data)  # Insert new data

# Example of how to use it in a function
def example_function():
    # Example of displaying some data in the right box
    show_data_in_right_box("This is some example data displayed on the right side of the window.")

# Add the "Match the Tracking Numbers" button to the GUI
match_tracking_button = tk.Button(root, text="Match the Tracking Numbers", command=match_tracking_numbers, bg=button_color, fg="white", font=("Arial", 12))
match_tracking_button.grid(row=5, column=0, columnspan=2, pady=10)

# Button to execute SQL query
execute_query_button = tk.Button(root, text="Execute SQL Query", command=execute_sql_query_thread, bg=button_color, fg="white", font=("Arial", 12))
execute_query_button.grid(row=3, column=1, padx=20, pady=10)

# Button for connecting to SQL
connect_sql_button = tk.Button(root, text="Connect to SQL", command=lambda: connect_to_sql(output_box, connection_status_label), image=connect_sql_icon, compound=tk.LEFT, bg=button_color, fg="white", font=("Arial", 12))
connect_sql_button.grid(row=0, column=1, padx=20, pady=10)

# Button to open an Excel file
open_file_button = tk.Button(root, text="Open Excel File", command=open_file, image=open_file_icon, compound=tk.LEFT, bg=button_color, fg="white", font=("Arial", 12))
open_file_button.grid(row=0, column=0, padx=20, pady=10)

# Run the Tkinter event loop
root.mainloop()
